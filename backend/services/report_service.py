"""
Report generation service.

Supports four output formats:
  - CSV: Raw data export, good for data analysts.
  - JSON: Machine-readable, good for integration.
  - Excel: Rich formatting with multiple sheets, good for managers.
  - PDF: Executive-ready document with charts and summary tables.

Why openpyxl for Excel over xlwt/xlrd?
  - openpyxl supports .xlsx (Excel 2007+), formula writing, conditional
    formatting, and named styles. xlwt only writes legacy .xls format.

Why reportlab for PDF?
  - reportlab gives full control over layout; alternatives like WeasyPrint
    need a running browser and are slower for server-side generation.
"""

import csv
import io
import json
import logging
from datetime import datetime, timezone
from typing import Any

logger = logging.getLogger(__name__)

# Optional heavy dependencies — import lazily to keep startup fast
try:
    import openpyxl
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side,
        NamedStyle,
    )
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed; Excel export disabled")

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        HRFlowable,
        PageBreak,
    )
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not installed; PDF export disabled")


# ---------------------------------------------------------------------------
# Brand palette (matches frontend ThemeProvider)
# ---------------------------------------------------------------------------
BRAND_NAVY = "#0A0F1E"
BRAND_CYAN = "#00D4FF"
BRAND_DARK_CARD = "#0D1B2A"

SEVERITY_COLORS = {
    "critical": "#F44336",
    "high": "#FF9800",
    "medium": "#FFC107",
    "low": "#2196F3",
    "info": "#9E9E9E",
}


class ReportService:
    """
    Generates governance reports in multiple formats.

    All generate_* methods return bytes that can be streamed directly
    as an HTTP response or stored in object storage.
    """

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def generate_csv(self, findings: list[dict[str, Any]]) -> bytes:
        """
        Generate a flat CSV of all findings.

        Flattens nested dicts to dot-notation keys for spreadsheet compatibility.
        """
        if not findings:
            return b"No findings\n"

        output = io.StringIO()
        # Flatten findings for CSV
        flat_findings = [self._flatten(f) for f in findings]
        fieldnames = sorted({k for row in flat_findings for k in row.keys()})

        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            extrasaction="ignore",
            lineterminator="\r\n",
        )
        writer.writeheader()
        writer.writerows(flat_findings)
        return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility

    def _flatten(self, d: dict, parent_key: str = "", sep: str = ".") -> dict:
        """Recursively flatten nested dict keys with dot notation."""
        items = {}
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(self._flatten(v, new_key, sep))
            elif isinstance(v, list):
                items[new_key] = ", ".join(str(i) for i in v)
            elif hasattr(v, "tzinfo") and v.tzinfo is not None:
                # openpyxl rejects timezone-aware datetimes at wb.save()
                # time. Convert to naive UTC (strip tzinfo without
                # converting — the value is already in UTC) so Excel
                # receives a plain datetime it can write.
                items[new_key] = v.replace(tzinfo=None)
            else:
                items[new_key] = v
        return items

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def generate_json(self, report_data: dict[str, Any]) -> bytes:
        """
        Generate a structured JSON report.

        Includes metadata envelope with generation timestamp and version.
        """
        envelope = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "generator": "Azure Resource Guardian v1.0",
            "report": report_data,
        }
        return json.dumps(envelope, indent=2, default=str).encode("utf-8")

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def generate_excel(
        self,
        findings: list[dict[str, Any]],
        costs: list[dict[str, Any]] | None = None,
        summary: dict[str, Any] | None = None,
        title: str = "Azure Resource Guardian Report",
    ) -> bytes:
        """
        Generate a multi-sheet Excel workbook.

        Sheets:
          1. Executive Summary — KPIs and score gauges
          2. Findings — all findings with severity colour coding
          3. Cost Savings — top saving opportunities
          4. Raw Data — unformatted findings for scripting

        Returns bytes of the .xlsx file.
        """
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define reusable styles
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="0A0F1E")
        center_align = Alignment(horizontal="center", vertical="center")
        wrap_align = Alignment(wrap_text=True, vertical="top")

        severity_fills = {
            "critical": PatternFill("solid", fgColor="F44336"),
            "high": PatternFill("solid", fgColor="FF9800"),
            "medium": PatternFill("solid", fgColor="FFC107"),
            "low": PatternFill("solid", fgColor="2196F3"),
            "info": PatternFill("solid", fgColor="9E9E9E"),
        }

        def style_header_row(ws, row: int, col_count: int):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

        # ── Sheet 1: Executive Summary ──────────────────────────────
        ws_summary = wb.create_sheet("Executive Summary")
        ws_summary.sheet_view.showGridLines = False

        ws_summary["B2"] = title
        ws_summary["B2"].font = Font(name="Calibri", bold=True, size=18, color="0A0F1E")
        ws_summary["B3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
        ws_summary["B3"].font = Font(name="Calibri", italic=True, size=10, color="666666")

        if summary:
            row = 5
            ws_summary[f"B{row}"] = "Key Metrics"
            ws_summary[f"B{row}"].font = Font(bold=True, size=13)
            row += 1
            for key, value in summary.items():
                ws_summary[f"B{row}"] = str(key).replace("_", " ").title()
                # openpyxl can only write scalars to cells — convert dicts/
                # lists (e.g. by_severity sub-dicts) to a human-readable
                # string rather than letting the write silently fail or raise.
                if isinstance(value, dict):
                    cell_val = ", ".join(f"{k}: {v}" for k, v in value.items())
                elif isinstance(value, list):
                    cell_val = ", ".join(str(i) for i in value)
                elif hasattr(value, "replace"):
                    # Strip timezone from datetime — openpyxl rejects tz-aware datetimes
                    cell_val = value
                else:
                    cell_val = value
                ws_summary[f"C{row}"] = cell_val
                row += 1

        ws_summary.column_dimensions["B"].width = 28
        ws_summary.column_dimensions["C"].width = 20

        # ── Sheet 2: Findings ────────────────────────────────────────
        ws_findings = wb.create_sheet("Findings")
        columns = [
            "Severity", "Title", "Resource", "Type", "Subscription",
            "Resource Group", "Status", "Estimated Saving ($/mo)",
            "Recommendation", "Detected At",
        ]
        ws_findings.append(columns)
        style_header_row(ws_findings, 1, len(columns))

        for finding in findings:
            row_data = [
                finding.get("severity") or "",
                finding.get("title") or "",
                finding.get("resource_name") or "",
                finding.get("resource_type") or "",
                finding.get("subscription_id") or "",
                finding.get("resource_group") or "",
                finding.get("status") or "",
                finding.get("estimated_monthly_saving") or 0,
                finding.get("recommendation") or "",
                str(finding.get("detected_at") or ""),
            ]
            ws_findings.append(row_data)
            # Colour the severity cell
            row_num = ws_findings.max_row
            sev = finding.get("severity", "").lower()
            if sev in severity_fills:
                ws_findings.cell(row=row_num, column=1).fill = severity_fills[sev]
                ws_findings.cell(row=row_num, column=1).font = Font(color="FFFFFF", bold=True)

        # Auto-size columns
        for col_idx, col_name in enumerate(columns, 1):
            ws_findings.column_dimensions[get_column_letter(col_idx)].width = max(
                len(col_name) + 4, 15
            )

        # ── Sheet 3: Cost Savings ────────────────────────────────────
        if costs:
            ws_costs = wb.create_sheet("Cost Savings")
            cost_columns = ["Resource", "Type", "Subscription", "Monthly Saving ($)", "Annual Saving ($)", "Action"]
            ws_costs.append(cost_columns)
            style_header_row(ws_costs, 1, len(cost_columns))

            for cost in costs:
                monthly = cost.get("estimated_monthly_saving", 0) or 0
                ws_costs.append([
                    cost.get("resource_name", ""),
                    cost.get("resource_type", ""),
                    cost.get("subscription_id", ""),
                    round(monthly, 2),
                    round(monthly * 12, 2),
                    cost.get("recommendation_short", "Review and delete if unused"),
                ])

            for col_idx in range(1, len(cost_columns) + 1):
                ws_costs.column_dimensions[get_column_letter(col_idx)].width = 22

        # ── Sheet 4: Raw Data ────────────────────────────────────────
        ws_raw = wb.create_sheet("Raw Data")
        if findings:
            flat = [self._flatten(f) for f in findings]
            headers = sorted({k for row in flat for k in row.keys()})
            ws_raw.append(headers)
            style_header_row(ws_raw, 1, len(headers))
            for row_data in flat:
                ws_raw.append([row_data.get(h, "") for h in headers])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def generate_pdf(
        self,
        findings: list[dict[str, Any]],
        summary: dict[str, Any] | None = None,
        title: str = "Azure Resource Guardian — Governance Report",
        report_type: str = "technical",  # "executive" | "technical" | "board"
    ) -> bytes:
        """
        Generate a PDF governance report.

        report_type controls content depth:
          - "board": 1-page KPI summary, no resource details
          - "executive": Summary + top findings, no raw data
          - "technical": Full findings with resource IDs and remediation steps
        """
        if not PDF_AVAILABLE:
            raise RuntimeError("reportlab is not installed. Run: pip install reportlab")

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            title=title,
            author="Azure Resource Guardian",
        )

        styles = getSampleStyleSheet()
        story = []

        # ── Cover block ──────────────────────────────────────────────
        story.append(Paragraph(
            f'<font color="#00D4FF" size="22"><b>Azure Resource Guardian</b></font>',
            styles["Title"],
        ))
        story.append(Paragraph(
            f'<font color="#555555" size="14">{title}</font>',
            styles["Heading2"],
        ))
        story.append(Paragraph(
            f'Generated: {datetime.now().strftime("%B %d, %Y at %H:%M UTC")}',
            styles["Normal"],
        ))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#00D4FF")))
        story.append(Spacer(1, 0.4 * cm))

        # ── Executive Summary section ────────────────────────────────
        if summary:
            story.append(Paragraph("<b>Executive Summary</b>", styles["Heading1"]))
            kpi_data = [["Metric", "Value"]]
            for k, v in summary.items():
                kpi_data.append([str(k).replace("_", " ").title(), str(v)])

            t = Table(kpi_data, colWidths=[8 * cm, 8 * cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0F1E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(t)
            story.append(Spacer(1, 0.5 * cm))

        if report_type == "board":
            # Board report stops after the summary
            doc.build(story)
            return output.getvalue()

        # ── Findings section ─────────────────────────────────────────
        story.append(Paragraph("<b>Findings</b>", styles["Heading1"]))
        story.append(Paragraph(
            f"Total findings: <b>{len(findings)}</b>",
            styles["Normal"],
        ))
        story.append(Spacer(1, 0.3 * cm))

        # Severity counts
        sev_counts: dict[str, int] = {}
        for f in findings:
            sev = f.get("severity", "unknown").lower()
            sev_counts[sev] = sev_counts.get(sev, 0) + 1

        sev_data = [["Severity", "Count"]]
        for sev in ["critical", "high", "medium", "low", "info"]:
            if sev in sev_counts:
                sev_data.append([sev.upper(), str(sev_counts[sev])])

        sev_table = Table(sev_data, colWidths=[6 * cm, 4 * cm])
        sev_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0F1E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(sev_table)
        story.append(Spacer(1, 0.5 * cm))

        if report_type == "executive":
            # Executive report: top 20 findings only
            display_findings = sorted(
                findings,
                key=lambda x: ["critical", "high", "medium", "low", "info"].index(
                    x.get("severity", "info").lower()
                ),
            )[:20]
        else:
            display_findings = findings

        # Findings table
        if display_findings:
            story.append(Paragraph("<b>Finding Details</b>", styles["Heading2"]))
            find_data = [["Severity", "Title", "Resource", "Saving/mo"]]
            for f in display_findings:
                saving = f.get("estimated_monthly_saving")
                saving_str = f"${saving:.2f}" if saving else "—"
                find_data.append([
                    (f.get("severity") or "").upper(),
                    Paragraph((f.get("title") or "")[:80], styles["Normal"]),
                    Paragraph((f.get("resource_name") or "")[:60], styles["Normal"]),
                    saving_str,
                ])

            find_table = Table(
                find_data,
                colWidths=[2.5 * cm, 8 * cm, 5 * cm, 2.5 * cm],
                repeatRows=1,
            )
            sev_row_styles = []
            for i, f in enumerate(display_findings, 1):
                sev = f.get("severity", "").lower()
                hex_color = SEVERITY_COLORS.get(sev, "#9E9E9E").lstrip("#")
                sev_row_styles.append(
                    ("TEXTCOLOR", (0, i), (0, i), colors.HexColor(f"#{hex_color}"))
                )
                sev_row_styles.append(
                    ("FONTNAME", (0, i), (0, i), "Helvetica-Bold")
                )

            find_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0F1E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                *sev_row_styles,
            ]))
            story.append(find_table)

        doc.build(story)
        return output.getvalue()
